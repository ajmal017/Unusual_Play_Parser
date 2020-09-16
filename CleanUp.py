import UnusualWhaleNet
import openpyxl
import json
import requests
import bs4 as bs
import os
import sqlite3
from UnusualWhaleNet import get_sector

BASE_DIR = os.getcwd() + "/UnusualWhales"  # Where the files for this program will be stored
CONFIG_DIR = BASE_DIR + "/config/"
FILE_DIR = BASE_DIR + "/historyByDate/"
DOWNLOAD_DIR = BASE_DIR + "/downloads/"
DEFAULT_DB = "WhalePlays.db"
DEFAULT_DB_LOCATION = CONFIG_DIR + DEFAULT_DB
KEYS = ["name", "sector"]


def bad_symbols_excel_to_json():
    """
    Creates a file in which to store the ticker symbols for the "bad" or no longer solvent companies

    Takes no parameters

    Returns no values
    """

    bad_dict = []
    need_sector_file = CONFIG_DIR + "son_xlsx/xlsx/66_etfs_need_sectors.xlsx"
    wb = openpyxl.load_workbook(need_sector_file)
    sheet = wb["Sheet1"]

    for i in range(2, sheet.max_row + 1):
        symbol = sheet["A%d" % i].value
        sector = ""

        if sheet["E%d" % i].value in ['X', 'x']:
            bad_dict.append(dict(zip(KEYS, (symbol, sector))))

        if sheet["E%d" % i].value in ["", None]:
            if sheet["F%d" % i].value in ["", None]:
                bad_dict.append(dict(zip(KEYS, (symbol, sector))))

    with open("bad_symbols.json", "a+") as file:
        file.write(json.dumps(bad_dict, indent = 4, sort_keys = True))


def good_symbols_excel_to_json():
    """
    Writes ticker names to json along with their sector from an excel file containing needing to be categorized symbols

    Parameters are none

    Returns no values
    """
    new_dict = []
    need_sector_file = CONFIG_DIR + "/json_xlsx/xlsx/66_etfs_need_sectors.xlsx"
    wb = openpyxl.load_workbook(need_sector_file)
    sheet = wb["Sheet1"]

    for i in range(2, sheet.max_row + 1):
        symbol = sheet["A%d" % i].value

        if sheet["E%d" % i].value not in ["", None, 'X', 'x']:
            sector = sheet["E%d" % i].value

        elif sheet["F%d" % i].value not in ["", None, 'X', 'x']:
            sector = sheet["F%d" % i].value

        else:
            continue

        new_dict.append(dict(zip(KEYS, (symbol, sector))))

        with open("got_sectors.json", "a+") as file:
            file.write(json.dumps(new_dict, indent = 4, sort_keys = True))


def update_symbol_name():
    # check bad ticker list, try with one less letter and try with a period before the last letter in case of match
    # xop1 to xop
    # rdsa to rds.a

    1 == 1


def gsa(which, symbol):
    """
    Was used to fill excel sheet with sectors parsed from etfchannel, etfdatabse and barcharts. ETFdb worked the best.

    Parameter which specifies which site to use. "bc" for barcharts, "ec" for etfchannek and "ed" for etfdatabase.
    Parameter symbol is the ticker symbol to reference

    Returns a list containing the sector(s)
    """

    ss = []
    bc_url = "https://www.barchart.com/etfs-funds/quotes/%s/overview"
    ec_url = "https://www.etfchannel.com/etfs/?symbol=%s"
    edb_url = "https://etfdb.com/etf/%s/#etf-ticker-profile"
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/531.36 (KHTML, like Gecko) Chrome/50.0.2661.101 Safari/537.56'}
    need_sector_file = CONFIG_DIR + "json_xlsx/xlsx/66_etfs_need_sectors.xlsx"
    need_sector = ['HYG', 'EFA', 'SPY', 'FXI', 'LQD', 'EWZ', 'EWJ', 'ERI', 'EEM', 'UVXY', 'VXX', 'TZA', 'SQQQ', 'IWM', 'TLT', 'TNA', 'SPXU', 'UPRO', 'SVXY', 'IWD',
                   'DIA', 'EWY', 'RDSA', 'EWW', 'AMLP', 'RSX', 'FEZ', 'IWF', 'CHK', 'EMB', 'IJR', 'ASHR', 'FCAU', 'SRTY', 'BKLN', 'LK', 'SPXS', 'VWO', 'VAL', 'BGG',
                   'SDS', 'LLEX', 'VIXY', 'TQQQ', 'GNC', 'SCHA', 'EWG', 'INDA', 'QID', 'EWH', 'YINN', 'VOO', 'SDOW', 'XOP1', 'PSQ', 'SPXL', 'JNK', 'MCHI', 'EWT',
                   'IEMG', 'OXY2', 'SH', 'AGG', 'VUG', 'IEF', 'XLC']

    if which == 'bc':
        r = requests.get(bc_url % symbol, headers = headers)
        soup = bs.BeautifulSoup(r.text, 'html.parser')

        try:
            s = soup.select("a[href*='/competitors?quoteSectors=']")
            [ss.append(i.text) for i in s]
            return ss

        except:
            return "N/A"

    elif which == 'ec':
        r = requests.get(ec_url % symbol, headers = headers)
        soup = bs.BeautifulSoup(r.text, 'html.parser')

        try:
            s = soup.select("a[href^='/lists/?a=category&c']")
            [ss.append(i.text) for i in s]
            return ss

        except:
            return "N/A"

    elif which == 'ed':
        r = requests.get(edb_url % symbol, headers = headers)
        soup = bs.BeautifulSoup(r.text, 'html.parser')

        try:
            s = soup.find("span", text = "Category")
            # print("Soup find:", s)
            s = s.find_next_sibling().text.replace('\n', '')
            # print("Find next sibling:", s)
            return s

        except:
            return "N/A"


def save_list_of_no_sector_symbols(hide_found = False):
    """
    Reads in an excel sheet with symbols unmatched to a sector. SAves results to local JSON.

    Parameter hide_found is used to suppress print statements for symbols that now have a matched sector.

    Returns a list a symbols that do not have a matched sector.
    """

    new_dict = []
    not_found_count = 0
    found_count = 0
    no_sector_list = []
    wb = openpyxl.load_workbook(CONFIG_DIR + "json_xlsx/xlsx/DB_ETFs_with_no_Sectors.xlsx")

    for i in range(wb["Sheet1"].max_row):
        # print(i)
        a = wb["Sheet1"]['A'][i].value
        # b = wb["Sheet1"]['B'][i].value

        b = get_sector(a, hide_found)

        if b is None:
            not_found_count += 1
            no_sector_list.append(a)

        else:
            found_count += 1
            new_dict.append(dict(zip(KEYS, (a, b))))

        # print(i, a, b)

    if input("View dict: ") == 'y':
        [print(i) for i in new_dict]

    if input("Write:") == 'y':
        with open(CONFIG_DIR + "json_xlsx/JSONS/found_their_sector.json", 'a+') as file:
            file.write(json.dumps(new_dict, indent = 4, sort_keys = True))

    print("Found:", found_count)
    print("Not found count:", not_found_count)

    wb.close()

    return no_sector_list


def save_found_sectors_to_json():
    """
    This function takes the ticker symbols from the db and save them to a local JSON file.
    This way api calls don't need to be made unless necessary.

    Takes no parameters

    Returns no values
    """

    conn = sqlite3.connect(CONFIG_DIR + "WhalePlays.db")
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT symbol, sector FROM Plays WHERE sector IS NOT \"ETF\" AND sector IS NOT \"\"")
    data = cur.fetchall()

    new_dict = []

    for d in data:
        new_dict.append(dict(zip(KEYS, d)))

    with open(BASE_DIR + "/UnusualWhales/config/json_xlsx/JSONS/stock_json.json", 'a+') as file:
        file.write(json.dumps(new_dict, indent = 4, sort_keys = True))

    conn.commit()
    cur.close()
    conn.close()


def parse_etfdb_files_for_sectors_symbols():
    """
    This function takes the sectors from the etfdb.com xlsx files and creates a sector json file filled with symbols for that sector

    Takes no parameters

    Returns no values
    """

    # This doesnt need to be called often, only when updating the excel files from ETFDB.com

    sectors = ['Materials',
               'Financials',
               'Technology',
               'Energy',
               'Healthcare',
               'Consumer Staples',
               'Utilities',
               'Industrials',
               'Consumer Discretionary',
               'Real Estate',
               'Communications']

    sector_base = CONFIG_DIR + "json_xlsx"

    for name in sectors:  # for every item in the list sectors
        os.chdir(sector_base + name)  # cd to the sector folder
        print(os.getcwd())  # print the sector folder

        new_dict = []

        for file in os.listdir():  # list the files in the folder
            if file != ".DS_Store" and not file.startswith("~$"):
                # print('\t' + file)

                wb = openpyxl.load_workbook(file)

                for cell_row in wb["Overview"]["A"][1:]:
                    if {"name": cell_row.value, "sector": name} not in new_dict:
                        new_dict.append(dict(zip(KEYS, (cell_row.value, name))))

                wb.close()

        with open(sector_base + "JSONS/" + name + ".json", 'a+') as json_file_writer:
            json_file_writer.write(json.dumps(new_dict, indent = 4, sort_keys = True))

    #        print('$' * 30)

    os.chdir(sector_base)
